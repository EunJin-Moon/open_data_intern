import time
import os

from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook

def solution(road_name):
    options = Options()
    # options.headless = True
    browser = webdriver.Chrome(executable_path="./chromedriver.exe", options=options)
    browser.get("http://d.forest.go.kr/")
    wait = WebDriverWait(browser, 20);

    # 로그인
    browser.find_element_by_css_selector("div.userinfo.type2").find_element_by_class_name("btn_log").click()
    time.sleep(1)
    browser.find_element_by_id("userid").send_keys("3610000")
    browser.find_element_by_id("password").send_keys("3610000")
    browser.find_element_by_class_name("btn_userlogin").click()

    # 가로수 시스템 진입
    browser.find_element_by_css_selector("nav#nav").find_element_by_css_selector("div.inner").find_element_by_css_selector("span.ico.mn4").click()
    browser.execute_script("fn_createMenuDiv('FEAS00020101','', '', '', true);")
    # tag = browser.find_element_by_css_selector("div._mCS_13").find_element_by_css_selector("ul.subsubmenu").find_element_by_css_selector("a").click()
    # time.sleep(2)
    wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="searchRoadList"]')))
    Select(browser.find_element_by_id("searchRoadList")).select_by_visible_text(road_name)
    browser.execute_script("fn_search();")
    time.sleep(5)

    page_size = int(browser.find_element_by_id("sp_1_paging").get_attribute('textContent'))
    print("페이지 사이즈 : ", page_size)
    print(type(page_size))

    #엑셀 데이터 입력
    load_wb = load_workbook(os.path.abspath("자동입력/" + road_name + ".xlsx"), data_only=True)
    load_ws = load_wb['Sheet1']

    all_values = []
    for idx, row in enumerate(load_ws.rows):
        if idx == 0:
            continue
        row_value = []
        for cell in row:
            row_value.append(cell.value)
        all_values.append(row_value)

    for idx, row in enumerate(all_values):
        if row[12] == '완료':
            continue
        if row[5] == None:
            continue
        wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="citySttreMainCO"]/div/div/button[1]'))).send_keys(Keys.PAGE_DOWN)
        complete = False
        print(row[3], "확인")
        time.sleep(3)
        element = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="paging_center"]/table/tbody/tr/td[4]/input')))
        element.clear()
        current_page = int(idx / 10) + 1
        print("페이지 추적 : " + str(current_page))
        element.send_keys(current_page)
        element.send_keys(Keys.ENTER)
        time.sleep(1)

        tree_keys = browser.find_elements_by_css_selector('td[aria-describedby="datalist_sigunDstntSttreNo"')
        memo = ''
        for tree_idx, tag in enumerate(tree_keys):
            print("나무번호", tag.get_attribute("title"), "입니다.")
            if tag.get_attribute("title") == str(row[3]):
                print(tag.get_attribute("title") + " : " + str(row[3]) + "매칭완료")
                browser.execute_script("fn_detail(" + str(tree_idx+1) + ");")
                #데이터 입력 로직
                #수종(나무별 조건 추가 필요)
                # time.sleep(3)
                wait.until(EC.element_to_be_clickable((By.ID, "koftrCd")))
                if row[4] == '은행' or row[4] == '느티' or row[4] == '향' or row[4] == '이팝':
                    tree_name = row[4] + '나무'
                elif row[4] == '메타':
                    tree_name = '메타세쿼이아'
                Select(browser.find_element_by_id("koftrCd")).select_by_visible_text(tree_name)
                
                #흉고직경
                wait.until(EC.presence_of_element_located((By.NAME, "wdptBhgdm")))
                if row[5] != None:
                    browser.find_element_by_name("wdptBhgdm").clear()
                    browser.find_element_by_name("wdptBhgdm").send_keys(str(row[5]))
                else:
                    memo = '측정불가'
                
                # 근원경값
                wait.until(EC.presence_of_element_located((By.NAME, "wdptStumpVal")))
                if row[5] != None:
                    browser.find_element_by_name("wdptStumpVal").clear()
                    browser.find_element_by_name("wdptStumpVal").send_keys(str(row[6]))
                else:
                    memo = '측정불가'

                #암수

                if row[7] != None and row[7].strip() == '암':
                    Select(browser.find_element_by_id("sttreGynndTpcd")).select_by_visible_text("암")
                elif row[7] != None and row[7].strip() == '수':
                    Select(browser.find_element_by_id("sttreGynndTpcd")).select_by_visible_text("수")
                else:
                    Select(browser.find_element_by_id("sttreGynndTpcd")).select_by_visible_text("미분류")
                
                
                #전선
                if row[8] != None:
                    Select(browser.find_element_by_id("eltwrFg")).select_by_value(row[8].strip())
                

                #보호틀
                if row[9] != None:
                    browser.find_element_by_name("prtctMldTpeCont").clear()
                    browser.find_element_by_name("prtctMldTpeCont").send_keys(row[9].strip())
                else:
                     browser.find_element_by_name("prtctMldTpeCont").clear()







                #현황사진
                # browser.execute_script("fnPopup5(1);")
                # file_path = "C:/python_workspace/등록사진"+road_name+"/"+row[9].strip()+".PNG"
                # browser.find_element_by_name("uploadBtn1").sendKeys(file_path).click()
                # browser.find_element_by_css_selector("//*[@id='uploadFile1']").click()//
                # Select(browser.find_element_by_xpath('//*[@id="uploadBtn1"]')).sendKeys(file_path)
                # browser.find_element_by_name("file1").sendKeys("C:/python_workspace/등록사진"+road_name+"/"+row[9].strip()+".PNG")
                # WebElement fileInput = driver.findElement(By.name("file1"))
                # browser.find_element_by_css_selector("input[type='file']").send_keys("C:/python_workspace/등록사진"+road_name+"/"+row[9].strip()+".PNG")
                # browser.find_element_by_name("fn_update_citySttrePictureInfo();").click()
                # browser.find_element_by_name("btn check").click()



                
                #가로내녹지유형
                if(len(row) >= 12 and row[11] != None):
                    Select(browser.find_element_by_xpath('//*[@id="stsrrGrnsTpeCd"]')).select_by_visible_text(row[11].strip())
                #등록
                browser.execute_script("fn_updateSttreM();")
                time.sleep(0.5)
                browser.find_element_by_css_selector('div.ui-dialog.ui-widget.ui-widget-content.ui-corner-all.ui-front.pagelayer.ui-draggable')\
                    .find_element_by_class_name("layer_foot")\
                    .find_element_by_css_selector("button.btn.check").click()
                time.sleep(3)
                browser.find_element_by_css_selector('div.ui-dialog.ui-widget.ui-widget-content.ui-corner-all.ui-front.pagelayer.ui-draggable')\
                    .find_element_by_class_name("layer_foot")\
                    .find_element_by_css_selector("button.btn.check").click()
                browser.execute_script("fn_list();")
                time.sleep(2)
                wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id="1"]/td[5]')))
                complete = True

                # 완료 처리
                load_ws.cell(idx+2, 13, '완료')
                load_wb.save("자동입력/" + road_name + ".xlsx")

                break